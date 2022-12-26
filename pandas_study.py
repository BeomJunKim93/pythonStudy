# -*- coding: utf-8 -*-
import openpyxl
import pandas as pd
import openpyxl as oxl
import numpy as np
import os
# try1
# def read_data(dir):
#     f = open(dir, 'r')
#     data = f.readlines()
#     f.close()
#     return data
#
# file_main = read_data("./5.북부여성발전센터.xlsx")
# try2
# wb = load_workbook('./pandas_study/5.북부여성발전센터.xlsx')
# ws = wb.active
#
# ws.unmerge_cells('A1:H1')
#
# df2 = pd.DataFrame(ws.unmerge_cells)
# print(df2)

# try 3
# print(wb.sheetnames)

# # 워크시트 불러오기
# ws = wb['기본정보']
#
# # 열 (Column) 값 확인하기
# for cell in ws['A']:
#     print(cell.value)
# print(len(ws['A']))

# try 4
# 엑셀 파일 불러오기
# file_main = './pandas_study/5.북부여성발전센터.xlsx'

# 워크시트 생성하기
wb = openpyxl.Workbook()
ws = wb.active

column_value = ['void_name','파일명','기관구분코드','책임자 E-mail','부서명','성명','연락처','E-mail','건물명','지번주소','도로명주소','준공년도','건축물층수','건축물구조','건축물방위','기준층 천장고','연면적','기준층 층고','지하주차장','PC수','근무인원','IP 발급수','방문인원','일 근무시간','근무마감 시간','수영장유무','수영장면적','증 개축여부','증.개축 변동면적','증․개축 년도','증.개축 변동층수','증.개축 용도변경결과','전기 계량기 번호','가스 계량기 번호']
print(len(column_value))

looprow = 1
for i in range(1, 34):
    ws.cell(row = looprow, column = i).value = column_value[i]

file_list = os.listdir('./pandas_study')

# 워크북 (파일 자체) 불러오기

iter_coord = [
    (3, 2),
    (3, 6),
    (6, 2),
    (6, 6),
    (7, 2),
    (7, 6),
    (10, 2),
    (11, 2),
    (12, 2),
    (13, 2),
    (13, 7),
    (14, 2),
    (15, 2),
    (15, 6),
    (16, 2),
    (17, 2),
    (17, 6),
    (18, 2),
    (18, 6),
    (19, 2),
    (19, 6),
    (20, 2),
    (23, 2),
    (23, 6),
    (26, 2),
    (26, 6),
    (27, 2),
    (27, 6),
    (28, 6),
    (32, 1),
    (32, 5)
]

key_values = ['기관구분코드', '책임자 E-mail', '부서명','성명','연락처','E-mail','건물명','지번주소','도로명주소','준공년도','건축물층수','건축물구조','건축물방위','기준층 천장고','연면적','지하주차장','PC수','근무인원','IP 발급수','방문인원','일 근무시간','근무마감 시간','수영장유무','수영장면적','증 개축여부','증.개축 변동면적','증․개축 년도','증.개축 변동층수','증.개축 용도변경결과','전기 계량기 번호','가스 계량기 번호']


#
# for i in range(0, len(file_list)):
#     test_dict[key_values[i]] = ws2.cell(row=iter_coord[i][0], column=iter_coold[i][1]).value

if len(key_values) != len(iter_coord):
    raise IOError('배열 길이가 맞지 않습니다.')


final_res = []

for file in file_list[0:3]:
    temp_dict = {}
    # 파일정보 설정
    wb2 = openpyxl.load_workbook('./pandas_study/%s' % file)
    ws2 = wb2['기본정보']

    # 데이터 로드
    temp_dict['파일명'] = file
    for i in range(len(iter_coord)):
        temp_dict[key_values[i]] = ws2.cell(row=iter_coord[i][0], column=iter_coord[i][1]).value
    #
    # for key in key_values:
    #     temp_dict[key] = ws2.cell(row=iter_coord[key_values.index(key)][0], column=iter_coord[key_values.index(key)][1]).value
    # test_dict['기관구분코드'] = ws2.cell(row=3, column=2).value  # 기관구분코드
    # test_dict['책임자 E-mail'] = ws2.cell(row=3, column=6).value  # 책임자 E-mail
    # ws.cell(row=looprow2, column=4).value = ws2.cell(row=6, column=2).value  # 부서명
    # ws.cell(row=looprow2, column=5).value = ws2.cell(row=6, column=6).value  # 성명
    # ws.cell(row=looprow2, column=6).value = ws2.cell(row=7, column=2).value  # 연락처
    # ws.cell(row=looprow2, column=7).value = ws2.cell(row=7, column=6).value  # E-mail
    # ws.cell(row=looprow2, column=8).value = ws2.cell(row=10, column=2).value  # 건물명
    # ws.cell(row=looprow2, column=9).value = ws2.cell(row=11, column=2).value  # 지번주소
    # ws.cell(row=looprow2, column=10).value = ws2.cell(row=12, column=2).value  # '도로명주소'
    # ws.cell(row=looprow2, column=11).value = ws2.cell(row=13, column=2).value  # ,'준공년도',
    # ws.cell(row=looprow2, column=12).value = ws2.cell(row=13, column=7).value  # '건축물층수', + '/' + ws2.cell(row = 13, column = 7)
    # ws.cell(row=looprow2, column=13).value = ws2.cell(row=14, column=2).value  # '건축물구조',
    # ws.cell(row=looprow2, column=14).value = ws2.cell(row=15, column=2).value  # '건축물방위',
    # ws.cell(row=looprow2, column=15).value = ws2.cell(row=15, column=6).value  # '기준층 천장고',
    # ws.cell(row=looprow2, column=16).value = ws2.cell(row=16, column=2).value  # '연면적',
    # ws.cell(row=looprow2, column=18).value = ws2.cell(row=17, column=2).value  # '지하주차장',
    # ws.cell(row=looprow2, column=19).value = ws2.cell(row=17, column=6).value  # 'PC수',
    # ws.cell(row=looprow2, column=20).value = ws2.cell(row=18, column=2).value  # '근무인원',
    # ws.cell(row=looprow2, column=21).value = ws2.cell(row=18, column=6).value  # 'IP 발급수',
    # ws.cell(row=looprow2, column=22).value = ws2.cell(row=19, column=2).value  # '방문인원',
    # ws.cell(row=looprow2, column=23).value = ws2.cell(row=19, column=6).value  # '일 근무시간',
    # ws.cell(row=looprow2, column=24).value = ws2.cell(row=20, column=2).value  # '근무마감 시간',
    # ws.cell(row=looprow2, column=25).value = ws2.cell(row=23, column=2).value  # '수영장유무',
    # ws.cell(row=looprow2, column=26).value = ws2.cell(row=23, column=6).value  # '수영장면적',
    # ws.cell(row=looprow2, column=27).value = ws2.cell(row=26, column=2).value  # '증 개축여부',
    # ws.cell(row=looprow2, column=28).value = ws2.cell(row=26, column=6).value  # '증.개축 변동면적',
    # ws.cell(row=looprow2, column=29).value = ws2.cell(row=27, column=2).value  # '증․개축 년도',
    # ws.cell(row=looprow2, column=30).value = ws2.cell(row=27, column=6).value  # '증.개축 변동층수',
    # ws.cell(row=looprow2, column=31).value = ws2.cell(row=28, column=6).value  # '증.개축 용도변경결과',
    # ws.cell(row=looprow2, column=32).value = ws2.cell(row=32, column=1).value  # '전기 계량기 번호',
    # ws.cell(row=looprow2, column=33).value = ws2.cell(row=32, column=5).value  # '가스 계량기 번호']
    # looprow2 = looprow2 + 1
    final_res.append(temp_dict)

print(final_res)
df_res = pd.DataFrame(final_res)
print(df_res)
print(df_res['전기 계량기 번호'])
raise IOError
# 저장
wb.save("./sample.xlsx")

print(wb)